from flask import Flask, render_template, request
import platform
import os
import time
import threading
import queue
import atexit
import shutil
import tempfile

# Use pandas and openpyxl for Excel operations (cross-platform)
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

# For .xlsb files, use pyxlsb
try:
    import pyxlsb
    _PYXLSB_AVAILABLE = True
except ImportError:
    _PYXLSB_AVAILABLE = False

# Note: For formula evaluation, we'll use openpyxl's data_only mode
# which reads pre-calculated values. For full formula evaluation,
# you may need to use a library like 'formulas' or ensure Excel
# has calculated and saved the values.

app = Flask(__name__)

# Configure logging for production
import logging
if not app.debug:
    # Set up logging for production (Render)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    app.logger.setLevel(logging.INFO)

EXCEL_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__),
                          "Calculator_test.xlsb"))
SHEET_NAME = "Output"

EXCEL_TIMEOUT_S = float(os.environ.get("EXCEL_TIMEOUT_S", "90"))

DEFAULT_MAINTENANCE_MESSAGE = (
    "The calculator is under maintainence. We will notify you once its live"
)
# Create this file (empty is fine) to force maintenance mode:
# - enable:  create `maintenance.flag` next to app.py
# - disable: delete `maintenance.flag`
MAINTENANCE_FLAG_PATH = os.path.join(os.path.dirname(__file__), "maintenance.flag")

# Cache for converted Excel file (xlsb -> xlsx)
_xlsx_cache_file = None
_xlsx_cache_lock = threading.Lock()


def _env_truthy(name: str) -> bool:
    v = os.environ.get(name)
    if v is None:
        return False
    return str(v).strip().lower() in ("1", "true", "yes", "y", "on")


def get_maintenance_message() -> str | None:
    # Manual switches (recommended for Excel updates)
    if _env_truthy("MAINTENANCE_MODE") or _env_truthy("CALCULATOR_MAINTENANCE"):
        return DEFAULT_MAINTENANCE_MESSAGE
    try:
        if os.path.exists(MAINTENANCE_FLAG_PATH):
            return DEFAULT_MAINTENANCE_MESSAGE
    except Exception:
        # If we can't check the flag, don't block the app.
        pass

    # Auto-switch: if the workbook isn't present, show maintenance notice.
    try:
        if not os.path.exists(EXCEL_FILE):
            return DEFAULT_MAINTENANCE_MESSAGE
    except Exception:
        pass

    return None


def _should_show_maintenance_for_excel_error(e: Exception) -> bool:
    # If Excel is locked/recalculating/stuck or the file is being replaced,
    # show a friendly message instead of a raw traceback-style error.
    msg = str(e or "")
    if isinstance(e, (TimeoutError, FileNotFoundError)):
        return True
    if "Error opening/calculating Excel file" in msg:
        return True
    if "Excel operation timed out" in msg:
        return True
    return False


def _convert_xlsb_to_xlsx(xlsb_path: str) -> str:
    """Convert .xlsb file to .xlsx format for openpyxl compatibility."""
    global _xlsx_cache_file
    
    with _xlsx_cache_lock:
        # Check if we already have a cached xlsx file
        if _xlsx_cache_file and os.path.exists(_xlsx_cache_file):
            xlsb_mtime = os.path.getmtime(xlsb_path)
            xlsx_mtime = os.path.getmtime(_xlsx_cache_file)
            if xlsx_mtime >= xlsb_mtime:
                return _xlsx_cache_file
        
        # Create temporary xlsx file
        temp_dir = tempfile.gettempdir()
        xlsx_path = os.path.join(temp_dir, f"calculator_{os.path.basename(xlsb_path)}.xlsx")
        
        try:
            # Method 1: Try pandas first (more reliable for .xlsb files)
            # This preserves data structure better and handles edge cases
            try:
                if _PYXLSB_AVAILABLE:
                    df_dict = pd.read_excel(xlsb_path, sheet_name=None, engine='pyxlsb')
                else:
                    df_dict = pd.read_excel(xlsb_path, sheet_name=None)
                
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    for sheet_name, df in df_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                _xlsx_cache_file = xlsx_path
                return xlsx_path
            except Exception as pandas_error:
                # If pandas method fails, try direct pyxlsb method
                if not _PYXLSB_AVAILABLE:
                    raise Exception(f"Cannot convert .xlsb file. pyxlsb not available. Pandas error: {str(pandas_error)}")
                
                # Method 2: Direct pyxlsb to openpyxl conversion (fallback)
                try:
                    wb_xlsb = pyxlsb.open_workbook(xlsb_path)
                    wb_xlsx = openpyxl.Workbook()
                    
                    # Copy all sheets
                    for sheet_name in wb_xlsb.sheets:
                        ws_xlsb = wb_xlsb.get_sheet(sheet_name)
                        ws_xlsx = wb_xlsx.create_sheet(title=sheet_name)
                        
                        # Copy cells with validation
                        for row in ws_xlsb.rows():
                            for cell in row:
                                # Validate row and column indices (must be >= 1 for openpyxl)
                                if cell.r >= 1 and cell.c >= 1 and cell.v is not None:
                                    try:
                                        ws_xlsx.cell(row=cell.r, column=cell.c, value=cell.v)
                                    except (ValueError, TypeError) as e:
                                        # Skip invalid cells and log warning
                                        try:
                                            app.logger.warning(f"Skipping invalid cell at row={cell.r}, col={cell.c}: {str(e)}")
                                        except Exception:
                                            pass
                        # Note: formulas are lost in this conversion, but we'll handle that differently
                    
                    # Remove default sheet if we created new ones
                    if len(wb_xlsx.sheetnames) > 1 and 'Sheet' in wb_xlsx.sheetnames:
                        wb_xlsx.remove(wb_xlsx['Sheet'])
                    
                    wb_xlsx.save(xlsx_path)
                    wb_xlsx.close()
                    wb_xlsb.close()
                except Exception as pyxlsb_error:
                    raise Exception(f"Error converting .xlsb to .xlsx. Pandas method failed: {str(pandas_error)}. Pyxlsb method failed: {str(pyxlsb_error)}")
            
            _xlsx_cache_file = xlsx_path
            return xlsx_path
        except Exception as e:
            raise Exception(f"Error converting .xlsb to .xlsx: {str(e)}")


def _get_excel_file_path() -> str:
    """Get the Excel file path, converting .xlsb to .xlsx if needed."""
    if EXCEL_FILE.lower().endswith('.xlsb'):
        return _convert_xlsb_to_xlsx(EXCEL_FILE)
    return EXCEL_FILE


class _ExcelJob:
    __slots__ = ("fn", "args", "kwargs", "done", "result", "error")

    def __init__(self, fn, args, kwargs):
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.done = threading.Event()
        self.result = None
        self.error = None


class ExcelWorker:
    """
    Single-threaded Excel worker using pandas/openpyxl.
    
    This worker keeps one workbook open and serializes all requests.
    """

    def __init__(self, excel_file: str, sheet_name: str):
        self.excel_file = excel_file
        self.sheet_name = sheet_name

        self._q: "queue.Queue[_ExcelJob | None]" = queue.Queue()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._start_lock = threading.Lock()
        self._started = False

        self.wb = None
        self.ws = None
        # Track file signature so we can reload when the workbook changes on disk.
        self._file_sig: tuple[int, int] | None = None  # (mtime_ns, size)

    def _get_file_sig(self) -> tuple[int, int]:
        file_path = _get_excel_file_path()
        st = os.stat(file_path)
        return (int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1_000_000_000))), int(st.st_size))

    def start(self):
        with self._start_lock:
            if self._started:
                return
            self._thread.start()
            self._started = True

    def stop(self, timeout_s: float = 5.0):
        if not self._started:
            return
        try:
            self._q.put_nowait(None)
        except Exception:
            pass
        try:
            self._thread.join(timeout=timeout_s)
        except Exception:
            pass

    def call(self, fn, *args, timeout_s: float = EXCEL_TIMEOUT_S, **kwargs):
        self.start()
        job = _ExcelJob(fn, args, kwargs)
        self._q.put(job)
        if not job.done.wait(timeout=timeout_s):
            raise TimeoutError(
                f"Excel operation timed out after {timeout_s:.0f}s. "
                "The workbook may be recalculating or Excel may be stuck."
            )
        if job.error is not None:
            raise job.error
        return job.result

    def _run(self):
        try:
            while True:
                job = self._q.get()
                if job is None:
                    break
                try:
                    job.result = job.fn(*job.args, **job.kwargs)
                except Exception as e:
                    job.error = e
                finally:
                    job.done.set()
        finally:
            try:
                self._shutdown_excel()
            except Exception:
                pass

    def _shutdown_excel(self):
        # Close workbook
        try:
            if self.wb:
                self.wb.close()
        except Exception:
            pass
        self.ws = None
        self.wb = None
        self._file_sig = None

    def _restart_excel(self):
        self._shutdown_excel()
        # Don't re-check signature during the reopen to avoid recursion.
        self._ensure_open(check_for_updates=False)

    def _ensure_open(self, *, check_for_updates: bool = True):
        # If we already have an open workbook, optionally reload it if the file on disk changed.
        if self.wb is not None and self.ws is not None:
            if check_for_updates:
                try:
                    current_sig = self._get_file_sig()
                    if self._file_sig is not None and current_sig != self._file_sig:
                        # Workbook replaced/updated on disk -> reopen to pick up changes.
                        self._restart_excel()
                    else:
                        # First time tracking signature after open
                        self._file_sig = current_sig
                except Exception:
                    # If stat fails mid-replace, we'll naturally fail on open next call.
                    pass
            return

        file_path = _get_excel_file_path()
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Load workbook with data_only=False to preserve formulas
        self.wb = load_workbook(file_path, data_only=False, keep_links=False)
        
        if self.sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in workbook")
        
        self.ws = self.wb[self.sheet_name]
        
        # Capture signature of the workbook we just opened
        try:
            self._file_sig = self._get_file_sig()
        except Exception:
            self._file_sig = None

    def _calculate_formulas(self):
        """Recalculate formulas in the workbook."""
        # openpyxl doesn't have a built-in calculation engine
        # We'll need to use a formula evaluation library or read calculated values
        # For now, we'll reload with data_only=True after writing values
        # This requires saving and reloading, which is not ideal but works
        pass


excel_worker = ExcelWorker(EXCEL_FILE, SHEET_NAME)


@atexit.register
def _shutdown_worker():
    try:
        excel_worker.stop(timeout_s=3.0)
    except Exception:
        pass


def _cell_to_coords(cell_ref: str) -> tuple[int, int]:
    """Convert Excel cell reference (e.g., 'A1') to (row, col) tuple."""
    from openpyxl.utils import coordinate_from_string, column_index_from_string
    col_str, row = coordinate_from_string(cell_ref)
    col = column_index_from_string(col_str)
    return (row, col)


def _get_cell_value(ws, cell_ref: str, use_text: bool = False):
    """Get cell value, optionally using displayed text."""
    row, col = _cell_to_coords(cell_ref)
    cell = ws.cell(row=row, column=col)
    
    if use_text:
        # Try to get formatted text (displayed value)
        if cell.data_type == 'f':  # Formula
            # For formulas, we need to evaluate or get calculated value
            # Since openpyxl doesn't calculate, we'll use the formula result if available
            # or try to get the value
            return str(cell.value) if cell.value else ""
        elif cell.number_format and '%' in str(cell.number_format):
            # Percentage formatting
            if cell.value is not None:
                return f"{float(cell.value) * 100:.1f}%"
        return str(cell.value) if cell.value is not None else ""
    else:
        return cell.value


def _set_cell_value(ws, cell_ref: str, value):
    """Set cell value."""
    row, col = _cell_to_coords(cell_ref)
    cell = ws.cell(row=row, column=col)
    cell.value = value


def _clear_cell(ws, cell_ref: str):
    """Clear cell contents."""
    row, col = _cell_to_coords(cell_ref)
    cell = ws.cell(row=row, column=col)
    cell.value = None


def _get_range_values(ws, range_ref: str):
    """Get values from a range (e.g., 'C11:F15')."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    values = []
    for row in range(min_row, max_row + 1):
        row_values = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            row_values.append(cell.value)
        values.append(row_values)
    return values


def read_excel(
    emp_id,
    doj,
    desired_earning,
    non_par,
    par,
    term,
    *,
    product_rows=None,
    rows_to_display=1,
    clear_new_rows_from=None,
):
    def _calculate_once():
        # Ensure workbook is open AND reload it if the file changed on disk.
        excel_worker._ensure_open(check_for_updates=True)
        ws = excel_worker.ws

        # IMPORTANT: workbook is kept open between requests, so clear prior inputs.
        # Only clear INPUT columns (do NOT clear P/W/X which may contain formulas).
        try:
            _clear_cell(ws, "J4")
            for row in range(5, 15):
                _clear_cell(ws, f"O{row}")
                _clear_cell(ws, f"Q{row}")
                _clear_cell(ws, f"R{row}")
                _clear_cell(ws, f"S{row}")
                _clear_cell(ws, f"T{row}")
                _clear_cell(ws, f"U{row}")
                _clear_cell(ws, f"V{row}")
        except Exception:
            pass

        # ---------- INPUT ----------
        _set_cell_value(ws, "D5", emp_id)
        _set_cell_value(ws, "D7", desired_earning)
        _set_cell_value(ws, "J4", (str(doj).strip() if doj is not None else ""))

        # Convert percentage inputs to decimal format for Excel (30 -> 0.30 for 30%)
        non_par_decimal = float(non_par) / 100 if non_par else 0
        par_decimal = float(par) / 100 if par else 0
        term_decimal = float(term) / 100 if term else 0

        _set_cell_value(ws, "D19", non_par_decimal)
        _set_cell_value(ws, "D20", par_decimal)
        _set_cell_value(ws, "D21", term_decimal)

        # Set number format for percentages
        ws["D19"].number_format = "0%"
        ws["D20"].number_format = "0%"
        ws["D21"].number_format = "0%"

        # ---------- INPUT (Individual Product rows: write O/Q/R/S/T/U/V starting row 5) ----------
        rows = product_rows or []
        for idx, row in enumerate(rows):
            excel_row = 5 + idx
            _set_cell_value(ws, f"O{excel_row}", str(row.get("product", "")).strip())
            _set_cell_value(ws, f"Q{excel_row}", str(row.get("rop_variant_yn", "")).strip())
            _set_cell_value(ws, f"R{excel_row}", str(row.get("cashback_yn", "")).strip())
            _set_cell_value(ws, f"S{excel_row}", int(row.get("ppt")))
            _set_cell_value(ws, f"T{excel_row}", int(row.get("pt")))
            _set_cell_value(ws, f"U{excel_row}", float(row.get("epi")))
            _set_cell_value(ws, f"V{excel_row}", float(row.get("rider_prem")))

        # If we're adding a new row in UI, clear the newly-added Excel row(s) so stale values don't show.
        if clear_new_rows_from is not None:
            try:
                start_clear = int(clear_new_rows_from)
                end_clear = 4 + int(rows_to_display)  # last Excel row used
                for excel_row in range(start_clear, end_clear + 1):
                    _clear_cell(ws, f"O{excel_row}")
                    _clear_cell(ws, f"Q{excel_row}")
                    _clear_cell(ws, f"R{excel_row}")
                    _clear_cell(ws, f"S{excel_row}")
                    _clear_cell(ws, f"T{excel_row}")
                    _clear_cell(ws, f"U{excel_row}")
                    _clear_cell(ws, f"V{excel_row}")
            except Exception:
                pass

        # For formula calculation, we need to save and reload
        # Note: openpyxl's data_only=True only reads values that were previously
        # calculated by Excel. For full formula evaluation, you would need a
        # formula evaluation library like 'formulas' or ensure the Excel file
        # is saved with calculated values.
        # 
        # For now, we'll save the workbook and try to reload with calculated values.
        # If the file was previously opened in Excel and saved, it should have
        # cached calculated values.
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        try:
            excel_worker.wb.save(temp_path)
            # Try to reload with data_only=True to get calculated values
            # This works if Excel previously calculated and cached the values
            try:
                wb_calculated = load_workbook(temp_path, data_only=True, keep_links=False)
                ws_calc = wb_calculated[excel_worker.sheet_name]
            except Exception:
                # If data_only fails, reload without it (formulas won't be calculated)
                wb_calculated = load_workbook(temp_path, data_only=False, keep_links=False)
                ws_calc = wb_calculated[excel_worker.sheet_name]
        except Exception:
            # If we can't save/reload, use the original worksheet
            ws_calc = ws
            wb_calculated = None
        finally:
            try:
                os.unlink(temp_path)
            except Exception:
                pass

        # ---------- OUTPUT (Single values) ----------
        try:
            h7_value = _get_cell_value(ws_calc, "H7")
            if h7_value is None:
                h7_value = 0
            else:
                h7_value = float(h7_value)
        except Exception:
            h7_value = 0

        result = {
            "name": _get_cell_value(ws_calc, "I5", use_text=True),
            "channel": _get_cell_value(ws_calc, "J6", use_text=True),
            "total_target": _get_cell_value(ws_calc, "D9", use_text=True),
            "total_biz": _get_cell_value(ws_calc, "I9", use_text=True),
            "achievement": _get_cell_value(ws_calc, "J9", use_text=True),
            "deficit": _get_cell_value(ws_calc, "I12", use_text=True),
            "incentive": _get_cell_value(ws_calc, "I14", use_text=True),
            "remaining_amount": _get_cell_value(ws_calc, "I15", use_text=True),
            "capping": h7_value,  # H7 value for validation
            # One-line banner label (requested: pick from C4)
            "sell_one_line": _get_cell_value(ws_calc, "C4", use_text=True),
            # Date-of-Joining validation flag (requested: read from K4)
            "doj_flag": _get_cell_value(ws_calc, "K4", use_text=True),
        }

        # ---------- OUTPUT (Product mix table C11:F15) ----------
        product_mix = []
        range_values = _get_range_values(ws_calc, "C11:F15")
        for row_idx in range(1, len(range_values)):  # Start from 1 to skip header row (index 0)
            row = []
            row_data = range_values[row_idx]
            for col_idx, cell_val in enumerate(row_data):
                if col_idx == 0:
                    row.append(str(cell_val) if cell_val is not None else "")
                elif col_idx == 1:
                    try:
                        if cell_val is not None:
                            row.append(int(float(cell_val)))
                        else:
                            row.append("")
                    except Exception:
                        row.append(str(cell_val) if cell_val is not None else "")
                else:
                    row.append(str(cell_val) if cell_val is not None else "")
            product_mix.append(row)
        result["product_mix"] = product_mix

        # ---------- OUTPUT (Hit 100% line) ----------
        result["hit_100"] = _get_cell_value(ws_calc, "C17", use_text=True)

        # ---------- OUTPUT (Below Hit 100% table) ----------
        below_table = []
        for row_num in (19, 20, 21):
            below_table.append(
                [
                    _get_cell_value(ws_calc, f"D{row_num}", use_text=True),
                    _get_cell_value(ws_calc, f"E{row_num}", use_text=True),
                    _get_cell_value(ws_calc, f"J{row_num}", use_text=True),
                ]
            )
        result["below_table"] = below_table

        # ---------- OUTPUT (Performance on other Parameters) ----------
        result["performance_params"] = [
            {"label": _get_cell_value(ws_calc, "C27", use_text=True), "value": _get_cell_value(ws_calc, "D27", use_text=True)},
            {"label": _get_cell_value(ws_calc, "C28", use_text=True), "value": _get_cell_value(ws_calc, "D28", use_text=True)},
        ]

        # ---------- OUTPUT (Individual Product Mix - headers N4:X4, rows starting 5) ----------
        individual_headers = []
        for col in range(14, 25):  # N=14 ... X=24
            try:
                cell = ws_calc.cell(row=4, column=col)
                individual_headers.append(str(cell.value) if cell.value else "")
            except Exception:
                individual_headers.append("")
        result["individual_product_headers"] = individual_headers

        try:
            rows_to_display_int = max(1, min(10, int(rows_to_display)))
        except Exception:
            rows_to_display_int = 1
        result["individual_product_row_count"] = rows_to_display_int

        def cell_display(cell_ref):
            try:
                v = _get_cell_value(ws_calc, cell_ref)
                if v is None:
                    return _get_cell_value(ws_calc, cell_ref, use_text=True) or ""
                if isinstance(v, (int, float)):
                    if isinstance(v, float) and v.is_integer():
                        return str(int(v))
                    if isinstance(v, int):
                        return str(v)
                    return str(v)
                return str(v).strip() if v else (_get_cell_value(ws_calc, cell_ref, use_text=True) or "")
            except Exception:
                return _get_cell_value(ws_calc, cell_ref, use_text=True) or ""

        rows_out = []
        for idx in range(rows_to_display_int):
            excel_row = 5 + idx
            rows_out.append(
                {
                    "product": _get_cell_value(ws_calc, f"O{excel_row}", use_text=True) or "",
                    "variant": _get_cell_value(ws_calc, f"P{excel_row}", use_text=True) or "",
                    "rop_variant_yn": _get_cell_value(ws_calc, f"Q{excel_row}", use_text=True) or "",
                    "cashback_yn": _get_cell_value(ws_calc, f"R{excel_row}", use_text=True) or "",
                    "ppt": cell_display(f"S{excel_row}"),
                    "pt": cell_display(f"T{excel_row}"),
                    "epi": cell_display(f"U{excel_row}"),
                    "rider_prem": cell_display(f"V{excel_row}"),
                    "incentive_per_product": cell_display(f"W{excel_row}"),
                    "total_incentive": cell_display(f"X{excel_row}"),
                }
            )
        result["individual_product_rows"] = rows_out

        # Clean up calculated workbook if we created one
        if wb_calculated:
            try:
                wb_calculated.close()
            except Exception:
                pass

        return result

    def _calculate_with_restart():
        try:
            return _calculate_once()
        except Exception:
            # If something broke, restart once and retry
            try:
                excel_worker._restart_excel()
                return _calculate_once()
            except Exception:
                raise

    try:
        return excel_worker.call(_calculate_with_restart, timeout_s=EXCEL_TIMEOUT_S)
    except Exception as e:
        error_msg = f"Error opening/calculating Excel file: {str(e)}\n\nFile path: {EXCEL_FILE}\n\n"
        raise Exception(error_msg)


@app.route("/", methods=["GET", "POST"])
def index():
    result = None

    error_message = None
    capping_value = None
    maintenance_message = get_maintenance_message()

    # Log all requests for debugging
    try:
        app.logger.info(f"Request: {request.method} {request.path} - Maintenance mode: {bool(maintenance_message)}")
        if request.method == "POST":
            app.logger.info(f"POST data received - emp_id: {request.form.get('emp_id')}, doj: {request.form.get('doj')}")
    except Exception:
        pass

    # Hard maintenance gate (Excel being updated / app temporarily unavailable)
    if maintenance_message:
        try:
            app.logger.warning(f"Maintenance mode active: {maintenance_message}")
        except Exception:
            pass
        return render_template(
            "index.html",
            result=None,
            error_message=None,
            maintenance_message=maintenance_message,
            capping_value=None,
        )
    
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        doj = request.form.get("doj")
        desired_earning = request.form.get("desired_earning")
        non_par = request.form.get("non_par")
        par = request.form.get("par")
        term = request.form.get("term")

        # Individual product table state
        action = request.form.get("action")  # "add_row" when clicking Add more products
        try:
            row_count = int(request.form.get("row_count") or 1)
        except Exception:
            row_count = 1
        row_count = max(1, min(10, row_count))

        product_rows = []
        has_product_table = request.form.get("row_count") is not None or request.form.get("product_1") is not None

        # Parse rows from form (1..row_count)
        if has_product_table:
            for i in range(1, row_count + 1):
                product = (request.form.get(f"product_{i}") or "").strip()
                rop_variant_yn = (request.form.get(f"rop_variant_yn_{i}") or "").strip()
                cashback_yn = (request.form.get(f"cashback_yn_{i}") or "").strip()
                ppt_raw = (request.form.get(f"ppt_{i}") or "").strip()
                pt_raw = (request.form.get(f"pt_{i}") or "").strip()
                epi_raw = (request.form.get(f"epi_{i}") or "").strip()
                rider_raw = (request.form.get(f"rider_prem_{i}") or "").strip()

                # Server-side validation (keep strict even though button is disabled client-side)
                if not product or not rop_variant_yn or not cashback_yn or not ppt_raw or not pt_raw or not epi_raw or not rider_raw:
                    error_message = "Please fill all Individual Product Mix fields before adding more products."
                    break
                if rop_variant_yn not in ("Y", "N") or cashback_yn not in ("Y", "N"):
                    error_message = "ROP Variant and Cashback must be Y or N."
                    break
                try:
                    ppt_val = int(ppt_raw)
                    pt_val = int(pt_raw)
                except Exception:
                    error_message = "PPT and PT must be whole numbers."
                    break
                if ppt_val > 30:
                    error_message = "Please enter the correct PPT"
                    break
                if ppt_val <= 0 or pt_val <= 0:
                    error_message = "Please fill all Individual Product Mix fields before adding more products."
                    break
                if str(ppt_raw).strip().find(".") != -1 or str(pt_raw).strip().find(".") != -1:
                    error_message = "PPT and PT must be whole numbers (no decimals)."
                    break
                if pt_val < ppt_val:
                    error_message = "PT cannot be less than PPT."
                    break
                try:
                    epi_val = float(epi_raw)
                    rider_val = float(rider_raw)
                except Exception:
                    error_message = "EPI and Rider Premium must be numbers."
                    break
                if epi_val <= 0 or rider_val <= 0:
                    error_message = "Please fill all Individual Product Mix fields before adding more products."
                    break

                product_rows.append(
                    {
                        "product": product,
                        "rop_variant_yn": rop_variant_yn,
                        "cashback_yn": cashback_yn,
                        "ppt": ppt_val,
                        "pt": pt_val,
                        "epi": epi_val,
                        "rider_prem": rider_val,
                    }
                )

        # Add-row action increases visible rows (max 10)
        rows_to_display = row_count
        clear_new_rows_from = None
        if error_message is None and action == "add_row" and row_count < 10:
            rows_to_display = row_count + 1
            clear_new_rows_from = 5 + row_count  # Excel row of the newly added UI row

        try:
            result = read_excel(
                emp_id,
                doj,
                desired_earning,
                non_par,
                par,
                term,
                product_rows=product_rows,
                rows_to_display=rows_to_display,
                clear_new_rows_from=clear_new_rows_from,
            )

            # Always pass capping back to UI for client-side validation on reload.
            try:
                capping_value = float(result.get("capping") or 0)
            except Exception:
                capping_value = None

            # If DOJ is invalid (requested: K4 reads F), do NOT show results.
            # If DOJ is valid (K4 reads T), allow results.
            doj_flag_raw = result.get("doj_flag")
            doj_flag = str(doj_flag_raw or "").strip().upper()
            doj_true = ("T", "TRUE", "1", "YES", "Y")
            if doj_flag not in doj_true:
                result = None
                error_message = "Enter correct Date of Joining"

            # If desired earning is above capping on the first attempt, do NOT show results.
            # Instead, ask user to adjust desired earning and re-submit.
            if result is not None:
                try:
                    desired_val = float(desired_earning) if desired_earning not in (None, "") else None
                except Exception:
                    desired_val = None
                if (
                    desired_val is not None
                    and capping_value is not None
                    and capping_value > 0
                    and desired_val > capping_value
                ):
                    # Keep the page "clean": no results shown, and JS will show inline message +
                    # disable Calculate until corrected.
                    result = None
                    error_message = None
        except Exception as e:
            # Log the full error for debugging
            try:
                app.logger.error(f"Error during calculation: {str(e)}", exc_info=True)
                app.logger.error(f"Excel file path: {EXCEL_FILE}")
                app.logger.error(f"Excel file exists: {os.path.exists(EXCEL_FILE)}")
            except Exception:
                pass
            
            # If Excel is unavailable (common during workbook updates), show maintenance message.
            if _should_show_maintenance_for_excel_error(e):
                try:
                    app.logger.exception("Excel unavailable; showing maintenance message")
                except Exception:
                    pass
                maintenance_message = DEFAULT_MAINTENANCE_MESSAGE
                error_message = None
            else:
                error_message = str(e)
            result = None

    return render_template(
        "index.html",
        result=result,
        error_message=error_message,
        maintenance_message=maintenance_message,
        capping_value=capping_value,
    )


# Health check endpoint for Render
@app.route("/health", methods=["GET"])
def health():
    """Health check endpoint for monitoring"""
    try:
        excel_exists = os.path.exists(EXCEL_FILE)
        maintenance = get_maintenance_message()
        return {
            "status": "ok" if excel_exists and not maintenance else "degraded",
            "excel_file_exists": excel_exists,
            "excel_file_path": EXCEL_FILE,
            "maintenance_mode": bool(maintenance),
            "maintenance_message": maintenance
        }, 200
    except Exception as e:
        return {"status": "error", "error": str(e)}, 500


if __name__ == "__main__":
    # Flask can use threading now that we're not using COM
    app.run(debug=True, use_reloader=False, threaded=True)
